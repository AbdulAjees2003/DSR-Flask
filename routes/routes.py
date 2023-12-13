from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_user, LoginManager, login_required, logout_user, current_user
from app import db
from models.sea_dsr import SEADSR
from models.air_dsr import AIRDSR
from models.login import USER
from werkzeug.security import generate_password_hash,check_password_hash
from logger import log_args_and_return
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


main_bp = Blueprint('main_bp', __name__)
login_manager = LoginManager()

@main_bp.route('/',methods=['GET', 'POST'])
# @main_bp.route('/login',methods=['GET', 'POST'])
# @log_args_and_return
def Login():
    try:    
        if request.method == 'POST':
            email = request.form.get('email')
            password = request.form.get('password')
            hashed_password = generate_password_hash(password)

            user = USER.query.filter_by(email=email).first()
            if user and check_password_hash(hashed_password,user.password):
                login_user(user)

                if user.accesstype == 'RW':
                    print("Redirecting to options route")
                    return redirect(url_for('main_bp.options'))
                elif user.accesstype == 'RO':
                    print("Redirecting to read_only_page route")
                    return redirect(url_for('main_bp.rooptions'))
            else:
                flash('Wrong Email & Password')
    except Exception as e:
        flash(f"An error occurred: {str(e)}")
    return render_template('login.html')


@login_manager.user_loader
def load_user(user_id):
    return USER.query.get(int(user_id))

# READ ONLY USERS

@main_bp.route('/rooptions')
# @log_args_and_return
@login_required
def rooptions():
    return render_template('RO Index.html')


@main_bp.route('/roair')
@login_required
def roair():
    # This route is only accessible to users with 'RO' access type
    # airinfo = AIRDSR.query.all()  

     # Get the company name from the currently logged-in user
    company_name = current_user.company

    # Filter AIRDSR records based on the company name
    airinfo = AIRDSR.query.filter_by(consignee=company_name)
    
    return render_template('RO Air Homepage.html',airinformation=airinfo)


@main_bp.route('/rosea')
@login_required
def rosea():
    # seainfo = SEADSR.query.all()

     # Get the company name from the currently logged-in user
    company_name = current_user.company

    # Filter AIRDSR records based on the company name
    seainfo = SEADSR.query.filter_by(consignee=company_name)

    # This route is only accessible to users with 'RO' access type
    return render_template('RO Sea Homepage.html',seainformation=seainfo)


# Read Write Users

@main_bp.route('/options')
# @log_args_and_return
@login_required
def options():
    return render_template('index.html')


# @log_args_and_return
def add_or_update_record(model_class,user=None, record_id=None, **data):
    try:
        current_time = datetime.now()

        if record_id:
            record = model_class.query.get_or_404(record_id)
            for key, value in data.items():
                setattr(record, key, str(value).upper() if isinstance(value, str) else value)

            record.updateddate = current_time
            record.updatedby = user.email if user else "SOMEUSER"  # You can replace this with the actual user if available
        else:
            # Add default values for the missing arguments
            data['createddate'] = current_time
            data['updateddate'] = current_time
            data['createdby'] = user.email if user else "SOMEUSER"  # You can replace this with the actual user if available
            data['updatedby'] = user.email if user else "SOMEUSER"  # You can replace this with the actual user if available

            record = model_class(**{k: str(v).upper() if isinstance(v, str) else v for k, v in data.items()})
            db.session.add(record)
        db.session.commit()
        # flash(f"DSR {'updated' if record_id else 'added'} successfully!")
    except Exception as e:
        flash(f"An error occurred: {str(e)}")


# AIR FUNCTIONS

@main_bp.route('/air',methods=['GET','POST'])
# @log_args_and_return
@login_required
def air():
    try:
        if request.method=="POST":
            # Fetch the latest serial number
            latest_sno = AIRDSR.query.order_by(AIRDSR.sno.desc()).first()

            # If there are existing records, increment the serial number; otherwise, start from 1
            new_sno = 1 if latest_sno is None else latest_sno.sno + 1

            data = {field: request.form[field].upper() for field in request.form if field != 'submit'}
            data['sno'] = new_sno  # Add the serial number to the data
            
            add_or_update_record(AIRDSR,user=current_user, **data)
            flash("Air DSR Submitted Successfully!")
            return redirect(url_for('main_bp.air'))
        return render_template('air_page.html')
    except Exception as e:
        flash(f"An error occurred: {str(e)}")
        return redirect(url_for('main_bp.air'))
    

@main_bp.route('/dsrairinfo', methods=['GET', 'POST'])
# @log_args_and_return
@login_required
def dsrairinfo():
    airinfo = AIRDSR.query.all()
    return render_template('Air Homepage.html',airinformation=airinfo)



@main_bp.route('/airupdate',methods=['GET','POST'])
# @log_args_and_return
@login_required
def airupdate():
    if request.method=="POST":
        air_form=AIRDSR.query.get(request.form.get('id'))
        data = {field: request.form[field].upper() for field in request.form if field != 'id'}
        add_or_update_record(AIRDSR,user=current_user, record_id=air_form.id, **data)
        flash(f"{air_form.ponumber} Has Updated Successfully")
        return redirect(url_for('main_bp.dsrairinfo'))


@main_bp.route('/airdelete/<id>',methods=['GET','POST'])
# @log_args_and_return
@login_required
def airdelete(id):
    try:
        air_form=AIRDSR.query.get_or_404(id)
        sno_to_update = air_form.sno
        db.session.delete(air_form)

        # Update sno for the remaining records
        remaining_records = AIRDSR.query.filter(AIRDSR.sno > sno_to_update).all()
        for record in remaining_records:
            record.sno -= 1
            
        db.session.commit()
        flash(f"{air_form.ponumber} Deleted Successfully")
    except Exception as e:
        flash(f"An error occurred while deleting: {str(e)}")
    return redirect(url_for('main_bp.dsrairinfo'))


@main_bp.route('/airdownload')
# @log_args_and_return
@login_required
def airdownload():
    try:
        # Specify the columns you want to include in the Excel file
        included_columns = ['sno', 'tsvbookingno', 'ponumber', 'suppliername', 'consignee', 'bookingreceiveddate', 'bkgdate','cargoreadiness','pickupdate','warehousercvd','countryoforigin','portofloading','countryofdestination','portofdischarge','terms','hawbno','mawbno','netwt','grswt','chgwt','pkgs','packagetype','commodity','flight1','flight2','flight3','etd','eta','revisedetd','revisedeta','prealertdtd','remarks']
        

        # Query data from the database
        data = AIRDSR.query.with_entities(*[getattr(AIRDSR, column) for column in included_columns]).all()

        # Create a DataFrame from the query result using only the included columns
        df = pd.DataFrame(data, columns=included_columns)

        # Change column headings to uppercase
        df.columns = [column.upper() for column in df.columns]

        # Change date format for date columns (assuming date columns end with 'date')
        date_columns = [col for col in df.columns if 'date' in col.lower()]
        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')  

        # Create an Excel file
        excel_filename = 'airdsr.xlsx'
        df.to_excel(excel_filename, index=False, engine='openpyxl')

        # Open the Excel file using openpyxl to modify the header color
        wb = Workbook()
        ws = wb.active
        for col_num, value in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=value)
            cell.fill = PatternFill(start_color='CFE0FE', end_color='CFE0FE', fill_type='solid')  # Set header background color

        # Append the data from the DataFrame to the worksheet
        for row_data in df.itertuples(index=False):
            ws.append(row_data)

        # Save the modified Excel file
        wb.save(excel_filename)

        # Return the Excel file as a response
        return send_file(excel_filename, as_attachment=True, download_name='airdsr.xlsx')

    except Exception as e:
        flash(f"An error occurred while downloading: {str(e)}")
        return redirect(url_for('main_bp.dsrairinfo'))


# SEA FUNCTIONS

@main_bp.route('/sea',methods=['GET','POST'])
# @log_args_and_return
@login_required
def sea():
    try:
        if request.method=="POST":
            # Fetch the latest serial number
            latest_sno = SEADSR.query.order_by(SEADSR.sno.desc()).first()

            # If there are existing records, increment the serial number; otherwise, start from 1
            new_sno = 1 if latest_sno is None else latest_sno.sno + 1

            data = {field: request.form[field].upper() for field in request.form if field != 'submit'}
            data['sno']=new_sno

            add_or_update_record(SEADSR,user=current_user, **data)
            flash("Sea DSR Submitted Successfully!")
            return redirect(url_for('main_bp.sea'))
        return render_template('sea_page.html')
    except Exception as e:
        flash(f"An error occurred: {str(e)}")
        return redirect(url_for('main_bp.sea'))


@main_bp.route('/dsrseainfo', methods=['GET', 'POST'])
# @log_args_and_return
@login_required
def dsrseainfo():
    seainfo = SEADSR.query.all()
    return render_template('Sea Homepage.html',seainformation=seainfo)


@main_bp.route('/seaupdate',methods=['GET','POST'])
# @log_args_and_return
@login_required
def seaupdate():
    if request.method=="POST":
        sea_form=SEADSR.query.get(request.form.get('id'))
        data = {field: request.form[field].upper() for field in request.form if field != 'id'}
        add_or_update_record(SEADSR,user=current_user, record_id=sea_form.id, **data)
        flash(f"{sea_form.ponumber} Has Updated Successfully")
        return redirect(url_for('main_bp.dsrseainfo'))


@main_bp.route('/seadelete/<id>',methods=['GET','POST'])
# @log_args_and_return
@login_required
def seadelete(id):
    try:
        sea_form=SEADSR.query.get_or_404(id)
        sno_to_update = sea_form.sno
        db.session.delete(sea_form)

        # Update sno for the remaining records
        remaining_records = SEADSR.query.filter(SEADSR.sno > sno_to_update).all()
        for record in remaining_records:
            record.sno -= 1

        db.session.commit()
        flash(f"{sea_form.ponumber} Deleted Successfully")
    except Exception as e:
        flash(f"An error occurred while deleting: {str(e)}")
    return redirect(url_for('main_bp.dsrseainfo'))



@main_bp.route('/seadownload')
# @log_args_and_return
@login_required
def seadownload():
    try:
        # Specify the columns you want to include in the Excel file
        included_columns = ['sno', 'tsvbookingno', 'ponumber', 'suppliername', 'consignee', 'materialpickupdate', 'actualpickupdate','etdcustomer','actualetd','etacustomer','actualeta','countryoforigin','portofloading','countryofdestination','portofdischarge','containertype','containertype1','containertype2','noofcontainers','containernumbers','netwt','grswt','chgwt','pkgs','packagetype','commodity','mblnumber','hblnumber','shippingliner','vesselvoyage','firstfeedername','firstfeederimono','transhipment1steta','transhipment1stetd','transhipment1stportname','mothervesselname','mothervesselimono','mothervesseleta','mothervesseletd','mothervesselportname','secondfeedername','secondfeederimono','transhipment2ndeta','transhipment2ndetd','transhipment2ndportname','thirdfeedername','thirdfeederimono','transhipment3rdeta','transhipment3rdetd','transhipment3rdportname','prealertdtd','remarks']

        # Query data from the database
        data = SEADSR.query.with_entities(*[getattr(SEADSR, column) for column in included_columns]).all()

        # Create a DataFrame from the query result using only the included columns
        df = pd.DataFrame(data, columns=included_columns)

        # Change column headings to uppercase
        df.columns = [column.upper() for column in df.columns]

        # Change date format for date columns (assuming date columns end with 'date')
        date_columns = [col for col in df.columns if 'date' in col.lower()]
        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')

        # Create an Excel file
        excel_filename = 'seadsr.xlsx'
        df.to_excel(excel_filename, index=False, engine='openpyxl')

        # Open the Excel file using openpyxl to modify the header color
        wb = Workbook()
        ws = wb.active
        for col_num, value in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=value)
            cell.fill = PatternFill(start_color='CFE0FE', end_color='CFE0FE', fill_type='solid')  # Set header background color

        # Append the data from the DataFrame to the worksheet
        for row_data in df.itertuples(index=False):
            ws.append(row_data)

        # Save the modified Excel file
        wb.save(excel_filename)

        # Return the Excel file as a response
        return send_file(excel_filename, as_attachment=True, download_name='seadsr.xlsx')

    except Exception as e:
        flash(f"An error occurred while downloading: {str(e)}")
        return redirect(url_for('main_bp.dsrseainfo'))
    


@main_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.')
    return redirect(url_for('main_bp.Login'))